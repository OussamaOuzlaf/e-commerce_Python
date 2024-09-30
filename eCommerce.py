from tkinter import *
from tkinter import ttk
import datetime
import openpyxl
from openpyxl import Workbook


root = Tk()
root.geometry("950x552")
root.iconbitmap("img/icon.ico")
root.title("Market tools for building")
root.resizable(False, False)
now = datetime.datetime.now()
dateFromComputeur = now.strftime("%Y-%m-%d")

wb = Workbook()
ws = wb.active;
ws.title = "Customer";
ws["A1"] = "Full Name";
ws["B1"] = "Number";
ws["C1"] = "Adresse";
ws["D1"] = "Total";
ws["E1"] = "Date";
wb.save("Products.xlsx")

F1 = Frame(root, bg="silver", width=600, height="552")
F1.place(x=1, y=1)
title = Label(F1, bg="#5F7161", text="Project To sell Tools Building", fg="white",
              font=("Tajwal", 13, 'bold'), width=60, pady=5)
title.place(x=0, y=0)

menu = {
    0: ["A saw", 20],
    1: ["Cart", 20],
    2: ["A axe", 20],
    3: ["shovel", 20],
    4: ["hammer", 20],
    5: ["bucket", 20],
    6: ["helmet", 20],
    7: ["knife", 20],
    8: ["pliers", 20],
    9: ["pliers", 20],
    10: ["screwdriver", 20],
    11: ["screwdriver", 20],
}

def bill():
    global EName
    global EPhone
    global EAdresse
    global ETotal
    global EDate
    L_img.place(x=955, y=435)
    root.geometry("1205x552")
    F = Frame(root, width=250, height=430, bg="#5F7161", bd=2)
    F.place(x=955, y=1)
    name = Label(F, text="Name", fg="white", bg="#5F7161", font=("Tajwal", 13, "bold"))
    name.place(x=5, y=5)
    EName = Entry(F, width=20, font=("Tajwal", 12, "bold"), justify=CENTER)
    EName.place(x=5, y=35)
    phone = Label(F, text="Phone", fg="white", bg="#5F7161", font=("Tajwal", 13, "bold"))
    phone.place(x=5, y=65)
    EPhone = Entry(F, width=20, font=("Tajwal", 12, "bold"), justify=CENTER)
    EPhone.place(x=5, y=95)
    adresse = Label(F, text="Adresse", fg="white", bg="#5F7161", font=("Tajwal", 13, "bold"))
    adresse.place(x=5, y=125)
    EAdresse = Entry(F, width=20, font=("Tajwal", 12, "bold"), justify=CENTER)
    EAdresse.place(x=5, y=155)
    total = Label(F, text="Total", fg="white", bg="#5F7161", font=("Tajwal", 13, "bold"))
    total.place(x=5, y=185)
    ETotal = Entry(F, width=20, font=("Tajwal", 12, "bold"), justify=CENTER)
    ETotal.place(x=5, y=215)
    date = Label(F, text="Date", fg="white", bg="#5F7161", font=("Tajwal", 13, "bold"))
    date.place(x=5, y=245)
    EDate = Entry(F, width=20, font=("Tajwal", 12, "bold"), justify=CENTER)
    EDate.place(x=5, y=275)
    B = Button(F, text="Save", width=29, bg="red", font=("Tajwal", 10, "bold"), fg="white", command=save)
    B.place(x=5, y=305)
    B = Button(F, text="Clear Spinbox", width=29, bg="red", font=("Tajwal", 10, "bold"), fg="white")
    B.place(x=5, y=335)
    B = Button(F, text="Search", width=29, bg="red", font=("Tajwal", 10, "bold"), fg="white")
    B.place(x=5, y=365)
    B = Button(F, text="Clear", width=29, bg="red", font=("Tajwal", 10, "bold"), fg="white")
    B.place(x=5, y=395)
    total = 0;

    for item in ttv.get_children():
        ttv.delete(item)
    for i in range(len(sb)):
        if(int(sb[i].get())>0):
            price = int(sb[i].get())*menu[i][1]
            total = total+price;
            mystr = (str(menu[i][1]), str(sb[i].get()), str(price))
            ttv.insert("", 'end', iid=i, text=menu[i][0], values=mystr)
    ETotal.insert("1", str(total) + "$")
    EDate.insert("1", str(dateFromComputeur))

def clear():
    for item in ttv.get_children():
        ttv.delete(item)
    EName.delete("0", END)
    EPhone.delete("0", END)
    EAdresse.delete("0", END)
    ETotal.delete("0", END)
    EDate.delete("0", END)

def save():
    name = EName.get();
    phone = EPhone.get();
    adresse = EAdresse.get();
    total = ETotal.get();
    date = EDate.get();
    excel = openpyxl.load_workbook("Products.xlsx");
    file = excel.active;
    file.cell(column=1, row=file.max_row+1, value=name)
    file.cell(column=2, row=file.max_row, value=phone)
    file.cell(column=3, row=file.max_row, value=adresse)
    file.cell(column=4, row=file.max_row, value=total)
    file.cell(column=5, row=file.max_row, value=date)
    excel.save("Products.xlsx")


img1 = PhotoImage(file="img/1.png")
img2 = PhotoImage(file="img/2.png")
img3 = PhotoImage(file="img/3.png")
img4 = PhotoImage(file="img/4.png")
img5 = PhotoImage(file="img/5.png")
img6 = PhotoImage(file="img/6.png")
img7 = PhotoImage(file="img/7.png")
img8 = PhotoImage(file="img/8.png")
img9 = PhotoImage(file="img/9.png")
img10 = PhotoImage(file="img/10.png")
img11= PhotoImage(file="img/11.png")
img12 = PhotoImage(file="img/12.png")

btn1 = Button(F1, bg="#EFEAD8", bd=1, relief=SOLID,
              cursor="hand2", width=88, height=85, image=img1, text="A saw", compound=TOP)
btn1.place(x=30, y=45)
btn2 = Button(F1, bg="#EFEAD8", bd=1, relief=SOLID,
              cursor="hand2", width=88, height=85, image=img2, text="Cart", compound=TOP)
btn2.place(x=170, y=45)
btn3 = Button(F1, bg="#EFEAD8", bd=1, relief=SOLID,
              cursor="hand2", width=88, height=85, image=img3, text="axe", compound=TOP)
btn3.place(x=310, y=45)
btn4 = Button(F1, bg="#EFEAD8", bd=1, relief=SOLID,
              cursor="hand2", width=88, height=85, image=img4, text="shovel", compound=TOP)
btn4.place(x=450, y=45)
btn5 = Button(F1, bg="#EFEAD8", bd=1, relief=SOLID,
              cursor="hand2", width=88, height=85, image=img5, text="hammer", compound=TOP)
btn5.place(x=30, y=170)
btn6 = Button(F1, bg="#EFEAD8", bd=1, relief=SOLID,
              cursor="hand2", width=88, height=85, image=img6, text="bucket", compound=TOP)
btn6.place(x=170, y=170)
btn7 = Button(F1, bg="#EFEAD8", bd=1, relief=SOLID,
              cursor="hand2", width=88, height=85, image=img7, text="helmet", compound=TOP)
btn7.place(x=310, y=170)
btn8 = Button(F1, bg="#EFEAD8", bd=1, relief=SOLID,
              cursor="hand2", width=88, height=85, image=img8, text="knife", compound=TOP)
btn8.place(x=450, y=170)
btn9 = Button(F1, bg="#EFEAD8", bd=1, relief=SOLID,
              cursor="hand2", width=88, height=85, image=img9, text="pliers", compound=TOP)
btn9.place(x=30, y=300)
btn10 = Button(F1, bg="#EFEAD8", bd=1, relief=SOLID,
              cursor="hand2", width=88, height=85, image=img10, text="pliers", compound=TOP)
btn10.place(x=170, y=300)
btn11 = Button(F1, bg="#EFEAD8", bd=1, relief=SOLID,
              cursor="hand2", width=88, height=85, image=img11, text="screwdriver", compound=TOP)
btn11.place(x=310, y=300)
btn12 = Button(F1, bg="#EFEAD8", bd=1, relief=SOLID,
              cursor="hand2", width=88, height=85, image=img12, text="screwdriver", compound=TOP)
btn12.place(x=450, y=300)

sb = [];
print(sb)
font1 = ('Timer', 12, 'normal');
sv1 = IntVar()
sv2 = IntVar()
sv3 = IntVar()
sv4 = IntVar()
sv5 = IntVar()
sv6 = IntVar()
sv7 = IntVar()
sv8 = IntVar()
sv9 = IntVar()
sv10 = IntVar()
sv11 = IntVar()
sv12 = IntVar()

sb1 = Spinbox(F1, from_=0, to=5, font=font1, width=9, textvariable=sv1)
sb1.place(x=30, y=140)
sb.append(sb1)
sb2 = Spinbox(F1, from_=0, to=5, font=font1, width=9, textvariable=sv2)
sb2.place(x=170, y=140)
sb.append(sb2)
sb3 = Spinbox(F1, from_=0, to=5, font=font1, width=9, textvariable=sv3)
sb3.place(x=310, y=140)
sb.append(sb3)
sb4 = Spinbox(F1, from_=0, to=5, font=font1, width=9, textvariable=sv4)
sb4.place(x=450, y=140)
sb.append(sb4)
sb5 = Spinbox(F1, from_=0, to=5, font=font1, width=9, textvariable=sv5)
sb5.place(x=30, y=270)
sb.append(sb5)
sb6 = Spinbox(F1, from_=0, to=5, font=font1, width=9, textvariable=sv6)
sb6.place(x=170, y=270)
sb.append(sb6)
sb7 = Spinbox(F1, from_=0, to=5, font=font1, width=9, textvariable=sv7)
sb7.place(x=310, y=270)
sb.append(sb7)
sb8 = Spinbox(F1, from_=0, to=5, font=font1, width=9, textvariable=sv8)
sb8.place(x=450, y=270)
sb.append(sb8)
sb9 = Spinbox(F1, from_=0, to=5, font=font1, width=9, textvariable=sv9)
sb9.place(x=30, y=400)
sb.append(sb9)
sb10 = Spinbox(F1, from_=0, to=5, font=font1, width=9, textvariable=sv10)
sb10.place(x=170, y=400)
sb.append(sb10)
sb11 = Spinbox(F1, from_=0, to=5, font=font1, width=9, textvariable=sv11)
sb11.place(x=310, y=400)
sb.append(sb11)
sb12 = Spinbox(F1, from_=0, to=5, font=font1, width=9, textvariable=sv12)
sb12.place(x=450, y=400)
sb.append(sb12)

BBtn1 = Button(F1, width=12,text="Buy Tools",bd=2 ,font=('Tajwal', 13, 'bold') , height=1, bg="#5F7161", fg="white", command=bill)
BBtn1.place(x=30, y=500)
BBtn2 = Button(F1, width=12,text="New Task",bd=2 ,font=('Tajwal', 13, 'bold') , height=1, bg="#5F7161", fg="white", command=clear)
BBtn2.place(x=170, y=500)
BBtn3 = Button(F1, width=12,text="Buy Tools",bd=2 ,font=('Tajwal', 13, 'bold') , height=1, bg="#5F7161", fg="white")
BBtn3.place(x=310, y=500)
BBtn4 = Button(F1, width=12,text="Quit",bd=2 ,font=('Tajwal', 13, 'bold') , height=1, bg="#5F7161", fg="white", command=quit)
BBtn4.place(x=450, y=500)

F2 = Frame(root, bg="gray", width=345, height=552)
F2.place(x=605, y=1)

ttv = ttk.Treeview(F2, selectmode='browse')
ttv.place(x=1, y=1, width=345, height=552)
ttv["columns"] = ('1', '2', '3')
ttv.column("0", width=70, anchor='c')
ttv.column("1", width=50, anchor='c')
ttv.column("2", width=50, anchor='c')
ttv.column("3", width=60, anchor='c')
ttv.heading("#0", text="Products", anchor="c")
ttv.heading("#1", text="Price", anchor="c")
ttv.heading("#2", text="Number", anchor="c")
ttv.heading("#3", text="Total", anchor="c")

img_Logo = PhotoImage(file="img/logo.png")
L_img = Label(root, image=img_Logo)

root.mainloop()